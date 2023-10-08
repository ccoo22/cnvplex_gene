import glob
import itertools
import logging
import os
import re
import sys
import warnings

import openpyxl
import pysam
import xlsxwriter

warnings.filterwarnings('ignore')
logging.basicConfig(
    format='[%(asctime)s %(levelname)s] %(message)s',
    stream=sys.stdout,
    level=logging.INFO
    )
log = logging.getLogger(__name__)


# 解析gtf注释列，拆分成字典
def parse_gtf_anno(annotation_string):
    annotation_dict = {}
    for anno in annotation_string.split('";'):
        if not re.search('\w', anno):
            continue
        name, value = anno.strip().split(' "', 1)
        annotation_dict[name] = value
    return annotation_dict


# 解析gtf文件，提取目标转录本
def parse_gtf(gtf):
    print(f"解析GTF {gtf}")
    features = ['exon', 'CDS', 'start_codon', 'stop_codon']
    gtf_info = {}
    with open(gtf, 'r') as fh:
        for line in fh:
            if re.match('#', line):
                continue
            # 提取gtf一行中的信息
            chrom, source, feature, start, end, score, strand, phread, annotation_string = line.strip().split('\t')
            if chrom.startswith('NW') or chrom.startswith('NT'):
                continue
            start, end = int(start), int(end)
            annotation_dict = parse_gtf_anno(annotation_string)
            gene_id = annotation_dict['gene_id']
            transcript_id = annotation_dict['transcript_id']
            
            # SHOX 特殊处理
            if transcript_id == 'NM_000451.4_1':
                continue
            if transcript_id == '':
                continue
            mrna = transcript_id
            # mrna = transcript_id.split('.')[0]
            if mrna not in gtf_info:
                gtf_info[mrna] = {}
                
            # 读入转录本的exon/start_codon/stop_codon/CDS 信息
            if feature in features:
                if feature not in gtf_info[mrna]:
                    gtf_info[mrna][feature] = []
                feature_number = annotation_dict.get('exon_number', '.')
                gtf_info[mrna][feature].append({'chrom': chrom, 'start': start, 'end': end, 'strand': strand, 'feature': feature, 'feature_number': feature_number, 'gene_id': gene_id})
    return gtf_info

# 解析gtf文件，提取目标转录本
def parse_gtf2(gtfs):
    gtf_info = {}
    
    for gtf in gtfs:
        print(f"解析GTF {gtf}")
        features = ['exon', 'CDS', 'start_codon', 'stop_codon']
        with open(gtf, 'r') as fh:
            for line in fh:
                if re.match('#', line):
                    continue
                # 提取gtf一行中的信息
                chrom, source, feature, start, end, score, strand, phread, annotation_string = line.strip().split('\t')
                if chrom.startswith('NW') or chrom.startswith('NT'):
                    continue
                start, end = int(start), int(end)
                annotation_dict = parse_gtf_anno(annotation_string)
                gene_id = annotation_dict['gene_id']
                transcript_id = annotation_dict.get('transcript_id', '')

                # # SHOX 特殊处理
                # if transcript_id == 'NM_000451.4_1':
                #     continue
                if transcript_id == '':
                    continue
                mrna = transcript_id
                mrna_clean = transcript_id.split('.')[0]  # 无版本号的
                if mrna_clean not in gtf_info:
                    gtf_info[mrna_clean] = {}
                if mrna not in gtf_info[mrna_clean]:
                    gtf_info[mrna_clean][mrna] = {"source": gtf}
                
                # 读入新的文件时，如果旧的文件已经有结果，则删除旧的
                if gtf_info[mrna_clean][mrna]['source'] != gtf:
                    gtf_info[mrna_clean][mrna] = {"source": gtf}
 
                # 读入转录本的exon/start_codon/stop_codon/CDS 信息
                if feature in features:
                    if feature not in gtf_info[mrna_clean][mrna]:
                        gtf_info[mrna_clean][mrna][feature] = []
                    feature_number = annotation_dict.get('exon_number', '.')
                    gtf_info[mrna_clean][mrna][feature].append({'chrom': chrom, 'start': start, 'end': end, 'strand': strand, 'feature': feature, 'feature_number': feature_number, 'gene_id': gene_id, 'mrna_id': mrna})
    return gtf_info


def parse_mrna_gene(files):
    '''读入基因、转录本名称信息'''
    mrna_gene = {}
    for file in files:
        with open(file, 'r') as fh:
            for line in fh:
                if not re.search('\w', line):
                    continue
                # print(line)
                gene, mrna, cnv, mutation = line.strip().split('\t')
                gene = gene.strip().split(' （')[0].split('（')[0]
                # mrna = mrna.strip().split('.')[0]
                mrna = mrna.strip()
                if mrna in mrna_gene:
                    print("重复的mrna" + mrna)
                    sys.exit(0)
                mrna_gene[mrna] = gene
    return mrna_gene

def parse_mrna_gene_green(file):
    '''green做特殊读取'''
    mrna_gene = {}
    with open(file, 'r') as fh:
        for line in fh:
            if not re.search('\w', line):
                continue
            # print(line)
            gene, mrna, exon_info1, exon_info_fixed = line.strip().split('\t')
            if exon_info_fixed != '.':
                exon_info1 = exon_info_fixed
            if exon_info1 == 'skip':
                continue
            gene = gene.strip().split(' （')[0].split('（')[0]
            # mrna = mrna.strip().split('.')[0]
            mrna = mrna.strip()
     
            # 初始化
            mrna_gene[mrna] = {'gene': gene, 'exon': [], 'exon_candidate': [], 'intron': [], 'blue': False }
            # 要设计的区域
            for content in re.split('[,，]',exon_info1):
                if content.startswith('外显子'):
                    # exon
                    content = content.replace('外显子', '')
                    if re.search('-', content):
                        # 是多个外显子选择一个
                        start, end = content.split('-')
                        mrna_gene[mrna]['exon'].append(str(start))
                        for number in range(int(start) + 1, int(end) + 1):
                            mrna_gene[mrna]['exon_candidate'].append(str(number))
                    else:
                        # 所有设定的外显子都要设计
                        for number in content.split('、'):
                            if not re.search('\d', number):
                                continue
                            if not re.search('^\d+$', number):
                                print('[Error] 输入数据异常，不是纯数字：', line)
                                sys.exit()
                            mrna_gene[mrna]['exon'].append(number)
                elif content.startswith('内含子'):
                    content = content.replace('内含子', '')
                    for number in content.split('、'):
                        if not re.search('\d', number):
                            continue
                        if not re.search('^\d+$', number):
                            print('[Error] 输入数据异常，不是纯数字：', line)
                            sys.exit()
                        mrna_gene[mrna]['intron'].append(number)
    return mrna_gene

def parse_probe_info(blue, yellow, green):
    '''解析三个设计原始文件'''
    blue_data = parse_mrna_gene([blue])
    yellow_data = parse_mrna_gene([yellow])
    green_data = parse_mrna_gene_green(green)
    summary = {'blue': blue_data, 'yellow': yellow_data, 'green': green_data}
    return summary


def judge_green_to_blue(mrna_gene_info, mrna_prepare):
    '''判断green基因是否设定为蓝色'''
    '1.如果选定外显子个数大于基因外显子个数的30%，则按蓝色标记基因设计规则设计；'
    genes = []
    for mrna in mrna_prepare['green'].keys():
        design_exon_count = len(mrna_prepare['green'][mrna]['exon'])
        exon_count = len(mrna_gene_info[mrna]['exon'])
        perc = round(design_exon_count / exon_count, 4)
        if perc > 0.3:
            genes.append(mrna)
            # print(f'[notice] {mrna} （共 {exon_count} 个外显子）设计了 {design_exon_count} 个外显子，占转录本的 {perc}, 因此设定为蓝色')
            mrna_prepare['blue'][mrna] = {'green': mrna_prepare['green'][mrna]['gene']}  # 用dict类型存储，用于区分原始的blue
            mrna_prepare['green'][mrna]['blue'] = True  # 加上标签，用于排除
    if genes:
        print(f'共 {len(genes)} 个基因的外显子数量大于 30%，设定为蓝色规则')
            

def calculate_intron(exons):
    '''基于exons坐标，计算出intron区域'''
    exons = sorted(exons, key=lambda region: region['start'])
    introns = []
    if len(exons) > 1:
        if exons[0]['strand'] == '+':
            # plus 方向
            for index in range(len(exons) - 1):
                exon = exons[index]
                exon_next = exons[index + 1]
                intron = {'chrom': exon['chrom'], 'start': exon['end'] + 1, 'end': exon_next['start'] - 1, 'strand': exon['strand'], 'feature': 'intron', 'feature_number': str(index + 1), 'gene_id': exon['gene_id'], 'mrna_id': exon['mrna_id']}
                introns.append(intron)
        else:
            # minus方向
            exons.reverse()  # 坐标从大到小排列
            for index in range(len(exons) - 1):
                exon = exons[index]
                exon_next = exons[index + 1]
                intron = {'chrom': exon['chrom'], 'start': exon_next['end'] + 1, 'end': exon['start'] - 1, 'strand': exon['strand'], 'feature': 'intron', 'feature_number': str(index + 1), 'gene_id': exon['gene_id'], 'mrna_id': exon['mrna_id']}
                introns.append(intron)
    
    return introns

def calculate_UTR5(exons, start_codon):
    '''根据外显子区域，以及起始密码子区域，计算出UTR5区域'''
    if start_codon['strand'] == '+':
        utr5_start = 0
        utr5_end = start_codon['start'] - 1
    else:
        utr5_start = start_codon['end'] + 1
        utr5_end = 1000 * 1000 * 1000 * 1000  # 使它足够大
    
    # 寻找overlap
    utr5s = []
    for exon in exons:
        # 是否有交集
        overlap_start = max(utr5_start, exon['start'])
        overlap_end = min(utr5_end, exon['end'])
        if overlap_end >= overlap_start:
            # 是UTR5
            utr5 = {'chrom': exon['chrom'], 'start': overlap_start, 'end': overlap_end, 'strand': exon['strand'], 'feature': 'UTR5', 'feature_number': exon['feature_number'], 'gene_id': exon['gene_id'], 'mrna_id': exon['mrna_id']}
            utr5s.append(utr5)
    return utr5s

def calculate_NONE_UTR3(exons, stop_codon):
    '''根据外显子区域，以及终止密码子区域，计算出非UTR3区域'''
    if stop_codon['strand'] == '+':
        utr5_start = 0
        utr5_end = stop_codon['start'] - 1
    else:
        utr5_start = stop_codon['end'] + 1
        utr5_end = 1000 * 1000 * 1000 * 1000  # 使它足够大
    
    # 寻找overlap
    utr5s = []
    for exon in exons:
        # 是否有交集
        overlap_start = max(utr5_start, exon['start'])
        overlap_end = min(utr5_end, exon['end'])
        if overlap_end >= overlap_start:
            # 是UTR5
            utr5 = {'chrom': exon['chrom'], 'start': overlap_start, 'end': overlap_end, 'strand': exon['strand'], 'feature': 'exon', 'feature_number': exon['feature_number'], 'gene_id': exon['gene_id'], 'mrna_id': exon['mrna_id']}
            utr5s.append(utr5)
    return utr5s

def calculate_UTR3(exons, stop_codon):
    '''根据外显子区域，以及终止密码子区域，计算出UTR3区域'''
    if stop_codon['strand'] == '+':
        utr3_start = stop_codon['end'] + 1
        utr3_end = 1000 * 1000 * 1000 * 1000  # 使它足够大
    else:
        utr3_start = 0
        utr3_end = stop_codon['start'] - 1
    
    # 寻找overlap
    utr3s = []
    for exon in exons:
        # 是否有交集
        overlap_start = max(utr3_start, exon['start'])
        overlap_end = min(utr3_end, exon['end'])
        if overlap_end >= overlap_start:
            # 是UTR5
            utr3 = {'chrom': exon['chrom'], 'start': overlap_start, 'end': overlap_end, 'strand': exon['strand'], 'feature': 'UTR3', 'feature_number': exon['feature_number'], 'gene_id': exon['gene_id'], 'mrna_id': exon['mrna_id']}
            utr3s.append(utr3)
    return utr3s

def read_mrna_gene_info(file):
    '''读取基因结构数据'''
    mrna_gene_info = {}
    with open(file, 'r') as fh:
        heads = fh.readline().strip().split('\t')
        for line in fh:
            values = line.strip().split('\t')
            tmps = { heads[col]: values[col]  for col in range(0,len(values))}
            feature = tmps['feature']
            if tmps['mrna'] not in mrna_gene_info:
                mrna_gene_info[tmps['mrna']] = {'mrna': tmps['mrna']}
            if feature not in mrna_gene_info[tmps['mrna']]:
                mrna_gene_info[tmps['mrna']][feature] = []
            mrna_gene_info[tmps['mrna']][feature].append({'chrom': tmps['chrom'], 'start': int(tmps['start']), 'end': int(tmps['end']), 'strand': tmps['strand'], 'feature': tmps['feature'], 'feature_number': tmps['feature_number'], "name": tmps['feature'] + tmps['feature_number'], 'width': int(tmps['width']), "gene_id": tmps["gene_id"], "mrna_id": tmps["mrna_id"]})
    return mrna_gene_info


def find_closest_region(one_obj, others_obj, need = 1, name_prefix=""):
    '''找到与目标区域中心点最近的一个区域'''
    center = int((one_obj['start'] + one_obj['end']) / 2)
    distance = {}
    for index, other_obj in enumerate(others_obj):
        if one_obj['chrom'] != other_obj['chrom']:
            continue
        # print(center, min_distance, min_index, index, other_obj,)
        distance[index] = min(abs(center - other_obj['start']), abs(center - other_obj['end']))
    ordered_index = sorted(distance.keys(), key=lambda index: distance[index])
    if need == 1:
        return others_obj[ordered_index[0]]
    if need == 2:
        one = others_obj[ordered_index[0]]
        two = []
        if len(others_obj) > 1:
            for index in ordered_index[1:]:
                tmp = others_obj[ordered_index[index]]
                two.append(name_prefix + tmp['name'] + "," + tmp['chrom'] + ":" + str(tmp['start']) + "-" + str(tmp['end']))
        if two:
            two = ";".join(two)
        else:
            two = '.'
        return one, two




# 探针设计
def design_yellow(designed_region, mrna_gene_info, yellow):
    '''黄色标记基因探针设计'''
    print('黄色标记基因探针设计')
    for mrna in yellow.keys():
        gene = yellow[mrna]
        start_codon = mrna_gene_info[mrna]['start_codon'][0]
        stop_codon = mrna_gene_info[mrna]['stop_codon'][0]
        chrom = start_codon['chrom']
        middle = int((min(start_codon['start'], stop_codon['start']) + max(start_codon['end'], stop_codon['end'])) / 2)
        near, candidate = find_closest_region({'chrom': chrom, 'start': middle, 'end': middle}, mrna_gene_info[mrna]['CDS'], 2, gene + "-")
        # 记录
        name = gene + '-1'
        designed_region[name] = {
            'name': name,
            'gene': gene,
            'mrna': mrna,
            'chrom': chrom,
            'center': start_codon['start'],
            'start': start_codon['start'] - 1000,
            'end': start_codon['start'] + 1000,
            'type': 'yellow-start_codon',
            'candidate': '.'
        }
        name = gene + '-2'
        designed_region[name] = {
            'name': name,
            'gene': gene,
            'mrna': mrna,
            'chrom': chrom,
            'center': middle,
            'start': near['start'],
            'end': near['end'],
            'type': 'yellow-closestCDS',
            'candidate': candidate
        }
        name = gene + '-3'
        designed_region[name] = {
            'name': name,
            'gene': gene,
            'mrna': mrna,
            'chrom': chrom,
            'center': stop_codon['start'],
            'start': stop_codon['start'] - 1000,
            'end': stop_codon['start'] + 1000,
            'type': 'yellow-stop_codon',
            'candidate': '.'
        }
        # 补充资料
        for name in [gene + '-1', gene + '-2', gene +'-3']:
            designed_region[name]['width'] = designed_region[name]['end'] - designed_region[name]['start'] + 1
            designed_region[name]['need_probe'] = 1  # 区域只需要1个探针
            designed_region[name]['message'] = '.'
            designed_region[name]['cnvplex'] = name + "," + designed_region[name]['chrom'] + ":" + str(designed_region[name]['start']) + "-" + str(designed_region[name]['end'])

def design_green(designed_region, mrna_gene_info, green):
    '''绿色标记基因探针设计'''
    print('绿色标记基因探针设计')
    for mrna in green.keys():
        # 设计的外显子占比超过30%，改为了blue
        if green[mrna]['blue']:
            continue
          
        # mrna_gene[mrna] = {'gene': gene, 'exon': [], 'exon_candidate': [], 'intron': [] }
        gene = green[mrna]['gene']
        for exon_number in green[mrna]['exon']:
            name = gene + "-exon" + exon_number
            # 找到外显子编号对应的区域
            exon = None
            for exon_tmp in mrna_gene_info[mrna]['exon']:
                if exon_number == exon_tmp['feature_number']:
                    exon = exon_tmp
                    break
            # 候选exon
            candidates = []
            for exon_candidate in green[mrna]['exon_candidate']:
                for exon_tmp in mrna_gene_info[mrna]['exon']:
                    if exon_candidate == exon_tmp['feature_number']:
                        candidates.append(gene + "-" + exon_tmp['name'] + "," + exon_tmp['chrom'] + ":" + str(exon_tmp['start']) + "-" + str(exon_tmp['end']))
                        break
            candidate = ";".join(candidates) if candidates else '.'
            if not exon:
                print(f'[Error] 转录本 {mrna} 不存在外显子 {exon_number}')
                designed_region[name] = {
                    'name': name,
                    'gene': gene,
                    'mrna': mrna,
                    'chrom': mrna_gene_info[mrna]['exon'][0]['chrom'],
                    'center': '.',
                    'start': '.',
                    'end': '.',
                    'width': '.',
                    'type': 'green-exon',
                    'need_probe': 2,
                    'candidate': candidate,
                    'cnvplex': '.',
                    'message': 'error 外显子不存在，需要确认是否写错了'
                }
            else:
                designed_region[name] = {
                    'name': name,
                    'gene': gene,
                    'mrna': mrna,
                    'chrom': exon['chrom'],
                    'center': int(exon['start'] + exon['end']),
                    'start': exon['start'],
                    'end': exon['end'],
                    'width': exon['end'] - exon['start'] + 1,
                    'type': 'green-exon',
                    'need_probe': 2,
                    'candidate': candidate,
                    'cnvplex': name + "," + exon['chrom'] + ":" + str(exon['start']) + "-" + str(exon['end']),
                    'message': '.'
                }
                
        for intron_number in green[mrna]['intron']:
            name = gene + "-intron" + intron_number
            # 找到外显子编号对应的区域
            intron = None
            for intron_tmp in mrna_gene_info[mrna]['intron']:
                if intron_number == intron_tmp['feature_number']:
                    intron = intron_tmp
                    break

            if not intron:
                print(f'[Error] 转录本 {mrna} 不存在内含子 {intron_number}')
                designed_region[name] = {
                    'name': name,
                    'gene': gene,
                    'mrna': mrna,
                    'chrom': mrna_gene_info[mrna]['intron'][0]['chrom'],
                    'center': '.',
                    'start': '.',
                    'end': '.',
                    'width': '.',
                    'type': 'green-intron',
                    'need_probe': 2,
                    'candidate': '.',
                    'cnvplex': '.',
                    'message': 'error 内含子不存在，需要确认是否写错了'
                }
            else:
                designed_region[name] = {
                    'name': name,
                    'gene': gene,
                    'mrna': mrna,
                    'chrom': intron['chrom'],
                    'center': int(intron['start'] + intron['end']),
                    'start': intron['start'],
                    'end': intron['end'],
                    'width': intron['end'] - intron['start'] + 1,
                    'type': 'green-intron',
                    'need_probe': 2,
                    'candidate': candidate,
                    'cnvplex': name + "," + intron['chrom'] + ":" + str(intron['start']) + "-" + str(intron['end']),
                    'message': '.'
                }

def design_blue(designed_region, mrna_gene_info, blue, genome):
    '''蓝色标记基因探针设计'''
    print('蓝色标记基因探针设计')
    # 1. 启动子区域设计 0.5K 1K 2K 上下扩展100
    print('    启动子区域设计')
    expand = 100
    for mrna in blue:
        gene = blue[mrna]
        source = 'blue'
        # 基因是否从green转移过来的
        if type(gene) == dict:
            gene = gene['green']
            source = 'green'
        
        # exon1
        exon1 = None
        for exon in mrna_gene_info[mrna]['exon']:
            if exon['feature_number'] == '1':
                exon1 = exon
                break

        positions = {
            '0.5K': exon1['start'] - 500,
            '1K': exon1['start'] - 1000,
            '2K': exon1['start'] - 2000,
        }
        if exon1['strand'] == '-':
            positions = {
                '0.5K': exon1['end'] + 500,
                '1K': exon1['end'] + 1000,
                '2K': exon1['end'] + 2000,
            }
        for dist in positions.keys():
            name = gene + "-promoter" + dist
            designed_region[name] = {
                    'name': name,
                    'gene': gene,
                    'mrna': mrna,
                    'gene_from': source,
                    'chrom': exon1['chrom'],
                    'center': positions[dist],
                    'start': positions[dist] - expand,
                    'end': positions[dist] + expand,
                    'width': expand * 2 + 1,
                    'type': 'blue-promoter',
                    'need_probe': 1,
                    'candidate': '.',
                    'message': '.'
                }
            designed_region[name]['cnvplex'] = name + "," + designed_region[name]['chrom'] + ":" + str(designed_region[name]['start']) + "-" + str(designed_region[name]['end'])
    
    # 2. UTR3 设计
    print('    UTR3设计')
    target_seq = 'AATAAA'
    target_seq_minus = "TTTATT"
    pysam_genome = pysam.FastaFile(genome) 
    for mrna in blue:
        gene = blue[mrna]
        source = 'blue'
        # 基因是否从green转移过来的
        if type(gene) == dict:
            gene = gene['green']
            source = 'green'
        stop_codon = mrna_gene_info[mrna]['stop_codon'][0]
        center = 0
        start = 0
        end = 0
        info = '.'
        for utr3_region in mrna_gene_info[mrna]['UTR3']:
            # 提取utr3序列
            seq = pysam_genome.fetch(region= utr3_region['chrom'] + ":" + str(utr3_region['start']) + "-" + str(utr3_region['end']))
            seq = seq.upper()
            # 确认是否存在目标序列
            match = None
            if utr3_region['strand'] == '+':
                match = re.search(target_seq, seq)
            else:
                match = re.search(target_seq_minus, seq)
            # 获取坐标
            if match:
                info = 'AATAAA序列'
                center = int((match.start() + match.end()) / 2) + utr3_region['start']
                start = match.start() + utr3_region['start'] - 50
                end = match.end() + utr3_region['start'] + 50 - 1
                break
        # 没有找到，则用转录终止点
        if info == '.':
            info = "转录终止点"
            if stop_codon['strand'] == '+':
                center = stop_codon['end']
                start = center - 100
                end = center
            else:
                center = stop_codon['start']
                start = center 
                end = center + 100
        # 记录
        name = gene + "-UTR3" 
        designed_region[name] = {
                    'name': name,
                    'gene': gene,
                    'mrna': mrna,
                    'gene_from': source,
                    'chrom': stop_codon['chrom'],
                    'center': center,
                    'start': start,
                    'end': end,
                    'width': end - start + 1,
                    'type': 'blue-utr3',
                    'need_probe': 1,
                    'candidate': '.',
                    'message': info
                }
        designed_region[name]['cnvplex'] = name + "," + designed_region[name]['chrom'] + ":" + str(designed_region[name]['start']) + "-" + str(designed_region[name]['end'])
    
    # 3. exon设计
    print('    exon区域设计（剔除UTR3区域）')
    for mrna in blue:
        gene = blue[mrna]
        source = 'blue'
        # 基因是否从green转移过来的
        if type(gene) == dict:
            gene = gene['green']
            source = 'green'
        stop_codon = mrna_gene_info[mrna]['stop_codon'][0]
        none_utr3s = calculate_NONE_UTR3(mrna_gene_info[mrna]['exon'], stop_codon)
        for region in none_utr3s:
            width = region['end'] - region['start'] + 1
            # 每1K 至少一个探针，最多4个探针
            probe_count = round(width / 1000)
            if probe_count == 0:
                probe_count = 1
            if probe_count > 4:
                probe_count = 4
            # 间隔
            interval = int(width / probe_count)
            # print(f"width: {width}  probe_count:{probe_count}   interval:{interval}")
            for index in range(probe_count):
                name = gene + "-exon" + region['feature_number'] + "-" + str(index + 1)
                start = region['start'] + index * interval
                end = start + interval - 1
                designed_region[name] = {
                    'name': name,
                    'gene': gene,
                    'mrna': mrna,
                    'gene_from': source,
                    'chrom': region['chrom'],
                    'center': int(start + end),
                    'start': start,
                    'end': end,
                    'width': end - start + 1,
                    'type': 'blue-exon',
                    'need_probe': 1,
                    'candidate': '.',
                    'message': f'当前外显子需要设计 {probe_count} 个探针'
                }
                designed_region[name]['cnvplex'] = name + "," + designed_region[name]['chrom'] + ":" + str(designed_region[name]['start']) + "-" + str(designed_region[name]['end'])



def read_probe(probe_dir):
    probe_info = {}
    duplicate_position = {}
    for subdir in glob.glob(probe_dir + "/*"):
        subname = os.path.basename(subdir)
        # 记录重复的探针位置，防止被多次使用
        for excel in glob.glob(f'{probe_dir}/{subname}/*.xlsx'):
            log.info(f'读入探针文件：{excel}')
            excel_name = os.path.basename(excel)
            wb = openpyxl.load_workbook(excel)
            ws_pair = wb['探针对信息']
            ws_probe = wb['探针信息']
            # （1）读入探针对信息
            # 表头
            heads = [ws_pair.cell(1, column).value for column in range(
                1, ws_pair.max_column + 1)]
            # 读入内容
            for row in range(2, ws_pair.max_row + 1):
                # 取出一行
                tmps = {}
                for column in range(1, ws_pair.max_column + 1):
                    tmps[heads[column - 1]
                         ] = ws_pair.cell(row, column).value
                # 保存
                if tmps['虚拟'] == 'TRUE' or tmps['虚拟'] == True:
                    continue
                probe_name =  subname + "-" + excel_name + '-' + tmps['探针对名称']
                chrom, start, end = tmps['染色体位置'].split('_')
                strand = tmps['探针方向'].split('/')[1]
                positions = [int(start), int(end)]
                positions.sort()

                # 检查是否有重复，重复的结果不用在保存了
                genome_position = '_'.join([chrom, str(positions[0]), str(positions[1]), strand])
                if genome_position in duplicate_position:
                    print("重复探针：", genome_position)
                    continue
                duplicate_position[genome_position] = 1
                probe_info[probe_name] = {}
                # probe_info[probe_name] = tmps
                probe_info[probe_name]['chrom'] = 'chr' + chrom
                probe_info[probe_name]['start'] = positions[0]
                probe_info[probe_name]['end'] = positions[1]
                probe_info[probe_name]['strand'] = strand
                probe_info[probe_name]['position'] = f"{chrom}:{start}-{end}"
                probe_info[probe_name]['probe_name'] = probe_name
                probe_info[probe_name]['最高同源性'] = round(float(tmps['最高同源性']), 2)
                probe_info[probe_name]['CNV%'] = tmps['CNV%']
                probe_info[probe_name]['SNP（MAF）'] = tmps['SNP（MAF）'] if tmps['SNP（MAF）'] != '' else '.'
                probe_info[probe_name]['used'] = False

            # (2) 读入探针序列信息
            heads = [ws_probe.cell(1, column).value for column in range(
                1, ws_probe.max_column + 1)]
            # 读入内容
            for row in range(2, ws_probe.max_row + 1):
                # 取出一行
                tmps = {}
                for column in range(1, ws_probe.max_column + 1):
                    tmps[heads[column - 1]
                         ] = ws_probe.cell(row, column).value
                # 保存
                if tmps['序列信息'] == '' or tmps['序列信息'] == None:
                    continue
                probe_name = subname + "-" + excel_name + '-' + tmps['对应探针对名称']
                # 可能因为坐标重复，而被删掉了
                if probe_name not in probe_info:
                    continue
                
                # 确定是 5‘ 3’
                probe_type_53 = re.match('\d', tmps['探针名称'].split('-')[-1]).group()  # 引物端号： 5 、 3
                probe_info[probe_name][f'{probe_type_53}_length'] = int(tmps['长度'])
                probe_info[probe_name][f'{probe_type_53}_tm'] = round(float(tmps['Tm']), 2)
                if probe_type_53 == '5':
                    probe_info[probe_name][f'5_seq'] = tmps['序列信息'].split(' ')[-1]
                if probe_type_53 == '3':
                    probe_info[probe_name][f'3_seq'] = tmps['序列信息'].split(' ')[0]
            wb.close()
    return probe_info

def workbook_format(workbook):
    """[给workbook添加几种常用的自定义样式，方便使用]

    Args:
        workbook ([obj]): [xlsxwriter.Workbook 对象]
    """
    wb_format = {}
    # 'font_size': 12,  # 字体大小
    # 'font_color': 'black',  # 字体颜色
    # 'font_name': 'Times New Roman',  # 字体
    # 'border': 1,  # 边框大小
    # 'border_color': 'black',  # 边框颜色
    # 'bg_color': 'yellow',  # 背景色
    # 'align': 'center',  # 水平对齐
    # 'valign': 'vcenter',  # 垂直对齐

    # 标题样式
    wb_format['title_style']  = workbook.add_format({'font_size': 12, 'font_color': 'black', 'font_name': 'Times New Roman', 'border': 1, 'border_color': 'black', 'bg_color': 'yellow', 'align': 'center', 'valign': 'vcenter',})
    # 常规样式
    wb_format['normal_style'] = workbook.add_format({'font_size': 12, 'font_color': 'black', 'font_name': 'Times New Roman', 'border': 1, 'border_color': 'black', 'bg_color': 'white', 'align': 'center', 'valign': 'vcenter',})
    wb_format['normal_style_left_align'] = workbook.add_format({'font_size': 12, 'font_color': 'black', 'font_name': 'Times New Roman', 'border': 1, 'border_color': 'black', 'bg_color': 'white', 'align': 'left', 'valign': 'vcenter',})
    wb_format['normal_style_mark'] = workbook.add_format({'font_size': 12, 'font_color': 'black', 'font_name': 'Times New Roman', 'border': 1, 'border_color': 'black', 'bg_color': '#c4d79b', 'align': 'center', 'valign': 'vcenter',})
    wb_format['normal_style_red'] = workbook.add_format({'font_size': 12, 'font_color': 'black', 'font_name': 'Times New Roman', 'border': 1, 'border_color': 'black', 'bg_color': 'red', 'align': 'center', 'valign': 'vcenter',})
    return wb_format

def output_blue(designed_region, workbook, wb_format):
    print("输出蓝色结果")
    # 探针信息
    probe_titles = ['chrom', 'start', 'end', 'strand', '5_seq', '5_length', '5_tm', '3_seq', '3_length', '3_tm']
    
    # excel表头
    titles = ['name', 'gene', 'mrna', 'gene_from', 'chrom', 'start', 'end', 'width', 'cnvplex', 'type', 'message']
    # excel表头追加
    for probe_title in probe_titles:
        titles.append('probe_' + probe_title)

    worksheet = workbook.add_worksheet('蓝色')
    row = 0
    for col in range(len(titles)):
        worksheet.write(row, col, titles[col], wb_format['title_style'])
    row += 1
    worksheet.freeze_panes(1, 0)  # 冻结第一行
    worksheet.set_column(0, 0, 20) # 修改列宽
    worksheet.set_column(2, 2, 15) # 修改列宽
    worksheet.set_column(5, 6, 15) # 修改列宽
    worksheet.set_column(10, 21, 12) # 修改列宽
    total = 0
    success = 0
    for name in sorted(designed_region.keys()):
        if not designed_region[name]['type'].startswith('blue'):
            continue
        total += 1
        if designed_region[name].get('probe_chrom', '.') != '.':
            success += 1
            
        values = [designed_region[name][title] if title in designed_region[name] else '.' for title in titles]
        
        for col in range(len(titles)):
            # 颜色标记
            style = 'normal_style'
            # 输出
            worksheet.write(row, col, values[col], wb_format[style])
        row += 1
    
    print(f"    共需要 {total} 个探针，成功设计出 {success} 个探针，完成率 {round(100 * success/total)} %， 缺失 {total - success} 个")
    

def output_yellow(designed_region, workbook, wb_format):
    print("输出黄色结果")
    # 探针信息
    probe_titles = ['chrom', 'start', 'end', 'strand', '5_seq', '5_length', '5_tm', '3_seq', '3_length', '3_tm']
    # excel表头
    titles = ['name', 'gene', 'mrna', 'chrom', 'start', 'end', 'center', 'width', 'cnvplex', 'type', 'candidate', 'message']
    # excel表头追加
    for probe_title in probe_titles:
        titles.append('probe_' + probe_title)

    worksheet = workbook.add_worksheet('黄色')
    row = 0
    for col in range(len(titles)):
        worksheet.write(row, col, titles[col], wb_format['title_style'])
    row += 1
    worksheet.freeze_panes(1, 0)  # 冻结第一行
    worksheet.set_column(0, 0, 20) # 修改列宽
    worksheet.set_column(2, 2, 15) # 修改列宽
    worksheet.set_column(4, 6, 15) # 修改列宽
    worksheet.set_column(10, 21, 12) # 修改列宽
    total = 0
    success = 0
    for name in sorted(designed_region.keys()):
        if not designed_region[name]['type'].startswith('yellow'):
            continue
        total += 1
        if designed_region[name].get('probe_chrom', '.') != '.':
            success += 1
        values = [designed_region[name][title] if title in designed_region[name] else '.' for title in titles]
        for col in range(len(titles)):
            # 颜色标记
            style = 'normal_style'
            # 输出
            worksheet.write(row, col, values[col], wb_format[style])
        row += 1
        
    print(f"    共需要 {total} 个探针，成功设计出 {success} 个探针，完成率 {round(100 * success/total)} %， 缺失 {total - success} 个")
    
def output_green(designed_region, workbook, wb_format):
    print("输出绿色结果")
    # 探针信息
    probe_titles = ['chrom', 'start', 'end', 'strand', '5_seq', '5_length', '5_tm', '3_seq', '3_length', '3_tm']
    # excel表头
    titles = ['name', 'gene', 'mrna', 'chrom', 'start', 'end', 'width', 'cnvplex', 'type', 'candidate', 'message', 'two_probe_distance', 'two_probe_overlap']
    # excel表头追加
    for probe_title in probe_titles:
        titles.append('probe1_' + probe_title)
    # excel表头追加
    for probe_title in probe_titles:
        titles.append('probe2_' + probe_title)
        
    worksheet = workbook.add_worksheet('绿色')
    row = 0
    for col in range(len(titles)):
        worksheet.write(row, col, titles[col], wb_format['title_style'])
    row += 1
    worksheet.freeze_panes(1, 0)  # 冻结第一行
    worksheet.set_column(0, 0, 20) # 修改列宽
    worksheet.set_column(2, 2, 15) # 修改列宽
    worksheet.set_column(4, 5, 15) # 修改列宽
    worksheet.set_column(10, 31, 12) # 修改列宽
    total = 0
    success = 0
    success_half = 0
    for name in sorted(designed_region.keys()):
        if not designed_region[name]['type'].startswith('green'):
            continue
        total += 1
        if designed_region[name].get('probe1_chrom', '.') != '.':
            if designed_region[name].get('probe2_chrom', '.') != '.':
                success += 1
            else:
                success_half += 1

        values = [designed_region[name][title] if title in designed_region[name] else '.' for title in titles]
        for col in range(len(titles)):
            # 颜色标记
            style = 'normal_style'
            # 输出
            worksheet.write(row, col, values[col], wb_format[style])
        row += 1
        
    
    print(f"    共需要 {total} 个探针，成功设计出 {success} 个探针，完成率 {round(100 * success/total)} %， 设计出一半的有  {success_half} 个， 缺失 {total - success - success_half}  + {success_half} 个")

def match_probe_by_region(designed_region, probe_list):
    print('对蓝色、黄色探针做区域匹配')
    probe_titles = ['chrom', 'start', 'end', 'strand', '5_seq', '5_length', '5_tm', '3_seq', '3_length', '3_tm']
    
    for name in designed_region.keys():
        # 只做蓝色、黄色
        if not designed_region[name]['type'].startswith('blue') and not designed_region[name]['type'].startswith('yellow'):
            continue
        # 当前探针可以设计的区域
        possible_regions = [{'chrom': designed_region[name]['chrom'], 'start': designed_region[name]['start'], 'end': designed_region[name]['end']}]
        if designed_region[name]['candidate'] != '.':
            for tmp in designed_region[name]['candidate'].split(';'):
                cds_name, region_info = tmp.split(',')
                chrom, start, end = re.split('[:-]', region_info)
                possible_regions.append({'chrom': chrom, 'start':int(start), 'end': int(end), 'name': cds_name})
                
        # 寻找区域内包含的探针
        for index, possible_region in enumerate(possible_regions):
            probe_info = {}
            for probe_name in probe_list.keys():
                if probe_list[probe_name]['used']:
                    continue
                # 计算重叠区域长度
                overlap_length = calculate_overlap_length(possible_region, probe_list[probe_name])
                if overlap_length == 0:
                    continue
                # 候选探针
                probe_info[probe_name] = max(abs(65 - probe_list[probe_name]['5_tm']), abs(65 - probe_list[probe_name]['3_tm']))
            # 选择tm最接近65的
            if probe_info:
                best_probe = sorted(probe_info.keys(), key=lambda probe_name: probe_info[probe_name])[0]
                # 储存探针信息
                for probe_title in probe_titles:
                    designed_region[name]['probe_' + probe_title] = probe_list[best_probe][probe_title]
                probe_list[best_probe]['used'] = True
                if index != 0:
                    designed_region[name]['message'] += f" 在候选区域 {possible_region['name']} 中设计"
                # 找到一个即可
                break


def calculate_overlap_length(region1, region2):
    '''计算两个区域的重叠长度'''
    overlap_length = 0
    if region1['chrom'] == region2['chrom']:
        overlap_start = max(region1['start'], region2['start'])
        overlap_end = min(region1['end'], region2['end'])
        if overlap_end >= overlap_start:
            overlap_length = overlap_end - overlap_start + 1
    return overlap_length


def match_probe_by_region2(designed_region, probe_list):
    print('对绿色探针做区域匹配')
    trantab = str.maketrans('ACGTacgt', 'TGCAtgca')
    probe_titles = ['chrom', 'start', 'end', 'strand', '5_seq', '5_length', '5_tm', '3_seq', '3_length', '3_tm']
    
    for name in designed_region.keys():
        # 只做蓝色、黄色
        if not designed_region[name]['type'].startswith('green'):
            continue
        # 有两个探针异常，不能做
        if designed_region[name]['start'] == '.':
            continue
        
        # 当前探针可以设计的区域
        possible_regions = [{'chrom': designed_region[name]['chrom'], 'start': designed_region[name]['start'], 'end': designed_region[name]['end']}]
        if designed_region[name]['candidate'] != '.':
            for tmp in designed_region[name]['candidate'].split(';'):
                cds_name, region_info = tmp.split(',')
                chrom, start, end = re.split('[:-]', region_info)
                possible_regions.append({'chrom': chrom, 'start':int(start), 'end': int(end), 'name': cds_name})
                
        for index, possible_region in enumerate(possible_regions):
            # 1. 寻找区域内包含的探针
            probe_info = {}
            for probe_name in probe_list.keys():
                if probe_list[probe_name]['used']:
                    continue
                # 计算重叠区域长度
                overlap_length = calculate_overlap_length(possible_region, probe_list[probe_name])
                if overlap_length == 0:
                    continue
                # 候选探针
                probe_info[probe_name] = max(abs(65 - probe_list[probe_name]['5_tm']), abs(65 - probe_list[probe_name]['3_tm']))

            # 2. 寻找区域内两两之间距离最远的探针
            if probe_info:
                probe_names = list(probe_info.keys())
                if len(probe_names) == 1:
                    # 2.1 区域内只设计出一个探针
                    best_probe = probe_names[0]
                    # 储存探针信息
                    for probe_title in probe_titles:
                        designed_region[name]['probe1_' + probe_title] = probe_list[best_probe][probe_title]
                    probe_list[best_probe]['used'] = True
                    designed_region[name]['message'] += f" 在候选区域 {possible_region['name']} 中设计了1个探针"
                else:
                    # 2.2 探针两两组合，计算距离，拿到距离最大的两个探针的编号
                    best_probe1 = '.'
                    best_probe2 = '.'
                    distance = 0
                    for probe_name1, probe_name2 in itertools.combinations(probe_names,2):
                        dist = abs(probe_list[probe_name1]['start'] - probe_list[probe_name2]['start'])
                        if dist > distance:
                            best_probe1 = probe_name1
                            best_probe2 = probe_name2
                            distance = dist

                    # 看一下两个探针的起始坐标，是否要调换一下顺序（从小到大）
                    if probe_list[best_probe1]['start'] > probe_list[best_probe2]['start']:
                        best_probe1, best_probe2 = best_probe2, best_probe1

                    # 储存探针信息
                    for probe_title in probe_titles:
                        designed_region[name]['probe1_' + probe_title] = probe_list[best_probe1][probe_title]
                    for probe_title in probe_titles:
                        designed_region[name]['probe2_' + probe_title] = probe_list[best_probe2][probe_title]
                    designed_region[name]['two_probe_distance'] = distance

                    # 两个探针之间是否重叠？
                    designed_region[name]['two_probe_overlap'] = calculate_overlap_length(probe_list[best_probe1], probe_list[best_probe2])

                    # 两个探针的strand方向是否一样？如果一样，修改第二个探针的方向
                    if probe_list[best_probe1]['strand'] == probe_list[best_probe2]['strand']:
                        # ['chrom', 'start', 'end', 'strand', '5_seq', '5_length', '5_tm', '3_seq', '3_length', '3_tm']
                        designed_region[name]['probe2_strand'] = '-' if probe_list[best_probe2]['strand'] == '+' else '+'
                        # 3/5 调换。同时序列反向互补
                        designed_region[name]['probe2_5_seq'] = probe_list[best_probe2]['3_seq'][::-1].translate(trantab) 
                        designed_region[name]['probe2_5_length'] = probe_list[best_probe2]['3_length']
                        designed_region[name]['probe2_5_tm'] = probe_list[best_probe2]['3_tm']

                        designed_region[name]['probe2_3_seq'] = probe_list[best_probe2]['5_seq'][::-1].translate(trantab) 
                        designed_region[name]['probe2_3_length'] = probe_list[best_probe2]['5_length']
                        designed_region[name]['probe2_3_tm'] = probe_list[best_probe2]['5_tm']

                    probe_list[best_probe1]['used'] = True
                    probe_list[best_probe2]['used'] = True
                    if index != 0:
                        designed_region[name]['message'] += f" 在候选区域 {possible_region['name']} 中设计"
                    break

                
def output_readme(workbook, wb_format):
    '''输出readme'''
        ##########
    # read me 表格
    ##########
    worksheet_readme = workbook.add_worksheet('read me')
    worksheet_readme.set_column(0, 0, 30)
    worksheet_readme.set_column(1, 1, 30)
    worksheet_readme.set_column(2, 2, 30)
    readme_info = [['sheet', 'title', 'description'],
                    ['蓝色/黄色/绿色', 'name', '区域命名。 蓝色： UTR3区域命名规则为 基因-UTR3； 外显子区域命名规则为 基因-exon外显子编号-区域编号，按照1K至少一个探针的设计要求，对exon区域进行等分，并编号1； promoter的 0.5K/1K/2K 命名为 基因-promoternK。黄色：基因-1 表示起始密码子范围； 基因-2 表示中间最近外显子；基因-3表示终止密码子。 绿色： 基因-exon外显子编号'],
                    ["蓝色/黄色/绿色", "gene", "基因"],
                    ["蓝色/黄色/绿色", "mrna", "转录本"],
                    ["蓝色/黄色/绿色", "chrom", "要设计探针的区域的染色体编号"],
                    ["蓝色/黄色/绿色", "start", "要设计探针的区域的起始坐标"],
                    ["蓝色/黄色/绿色", "end", "要设计探针的区域的终止坐标"],
                    ["蓝色/黄色/绿色", "width", "区域大小"],
                    ["蓝色/黄色/绿色", "type", "区域类型"],
                    ["蓝色/黄色/绿色", "message", "补充的提示信息"],
                    ["蓝色/黄色/绿色", "probe_name", "探针原始名称。从cnvplex软件提取出来的，格式： 文件名-探针名。仅供溯源参考"],
                    ["蓝色/黄色/绿色", "probe_chrom", "cnvplex软件设计结果：探针的染色体"],
                    ["蓝色/黄色/绿色", "probe_start", "cnvplex软件设计结果：探针的起始坐标"],
                    ["蓝色/黄色/绿色", "probe_end", "cnvplex软件设计结果：探针的终止坐标"],
                    ["蓝色/黄色/绿色", "probe_strand", "cnvplex软件设计结果：探针的方向"],
                    ["蓝色/黄色/绿色", "probe_5_seq", "cnvplex软件设计结果：5'探针序列"],
                    ["蓝色/黄色/绿色", "probe_5_length", "cnvplex软件设计结果：5'探针序列长度"],
                    ["蓝色/黄色/绿色", "probe_5_tm", "cnvplex软件设计结果：5'探针序列tm值"],
                    ["蓝色/黄色/绿色", "probe_3_seq", "cnvplex软件设计结果：3'探针序列"],
                    ["蓝色/黄色/绿色", "probe_3_length", "cnvplex软件设计结果：3'探针序列长度"],
                    ["蓝色/黄色/绿色", "probe_3_tm", "cnvplex软件设计结果：3'探针序列tm值"],
                    ["绿色", "two_probe_distance", "两个探针start坐标的距离"],
                    ["绿色", "two_probe_overlap", "两个探针的overlap长度"],
                    ["黄色/绿色", "candidate", "候选的可以设计探针的区域"],
                  ]
    row = 0
    for sheet, title, desc in readme_info:
        style = 'title_style' if row == 0 else 'normal_style_left_align'
        worksheet_readme.write(row, 0, sheet, wb_format[style])
        worksheet_readme.write(row, 1, title, wb_format[style])
        worksheet_readme.write(row, 2, desc, wb_format[style])
        row += 1